# archivo: core_admin/exports.py
"""
Exportación a Excel del histórico de consultas.

Genera un libro con dos hojas:
  1. "Consultas por mes"  → matriz meses x empresas, con totales y promedios.
  2. "Resumen por empresa" → una fila por empresa con su total y su promedio.

El rango de meses es continuo (desde la primera hasta la última consulta
registrada), incluyendo los meses sin actividad con valor 0. Esto hace que el
promedio mensual sea real y no quede inflado por los meses en que no se usó
la plataforma.
"""

from datetime import date

from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from consultas.models import Busqueda
from empresas.models import Empresa

# --- Paleta VADOM ---
COLOR_PRIMARIO = '1B7783'
COLOR_TOTALES = 'E8F3F4'
COLOR_PROMEDIO = 'FFF4E0'

MESES_ES = [
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
]

SIN_EMPRESA = 'Sin empresa asignada'

_BORDE = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def _nombre_mes(fecha):
    """'2026-03' → 'Marzo 2026'."""
    return f"{MESES_ES[fecha.month - 1].capitalize()} {fecha.year}"


def _rango_meses(inicio, fin):
    """Lista continua de fechas (día 1) desde el mes 'inicio' hasta 'fin'."""
    meses = []
    year, month = inicio.year, inicio.month
    while (year, month) <= (fin.year, fin.month):
        meses.append(date(year, month, 1))
        month += 1
        if month == 13:
            year, month = year + 1, 1
    return meses


def _recolectar_datos():
    """
    Agrupa todas las búsquedas por mes y empresa.

    Retorna (meses, empresas, conteos, total_general) donde:
      - meses: lista continua de fechas (día 1 de cada mes)
      - empresas: lista de nombres de empresa (columnas del reporte)
      - conteos: dict {(mes, empresa): cantidad}
    """
    filas = (
        Busqueda.objects
        .annotate(mes=TruncMonth('fecha_busqueda'))
        .values('mes', 'usuario__empresa__nombre')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    conteos = {}
    meses_con_datos = []
    hay_sin_empresa = False

    for fila in filas:
        # TruncMonth devuelve datetime con tz; lo pasamos a fecha local.
        mes_dt = fila['mes']
        if timezone.is_aware(mes_dt):
            mes_dt = timezone.localtime(mes_dt)
        mes = date(mes_dt.year, mes_dt.month, 1)

        nombre_empresa = fila['usuario__empresa__nombre'] or SIN_EMPRESA
        if nombre_empresa == SIN_EMPRESA:
            hay_sin_empresa = True

        conteos[(mes, nombre_empresa)] = conteos.get((mes, nombre_empresa), 0) + fila['total']
        meses_con_datos.append(mes)

    if not meses_con_datos:
        return [], [], {}, 0

    meses = _rango_meses(min(meses_con_datos), max(meses_con_datos))

    # Todas las empresas registradas (aunque no tengan consultas) + orfanas.
    empresas = list(Empresa.objects.order_by('nombre').values_list('nombre', flat=True))
    if hay_sin_empresa:
        empresas.append(SIN_EMPRESA)

    total_general = sum(conteos.values())
    return meses, empresas, conteos, total_general


def _estilar_encabezado(ws, fila, ultima_columna):
    for col in range(1, ultima_columna + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = Font(bold=True, color='FFFFFF', size=11)
        celda.fill = PatternFill('solid', fgColor=COLOR_PRIMARIO)
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border = _BORDE


def _ajustar_anchos(ws, anchos):
    for indice, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = ancho


def _hoja_por_mes(wb, meses, empresas, conteos):
    """Hoja 1: matriz meses x empresas + fila de totales y de promedio."""
    ws = wb.active
    ws.title = 'Consultas por mes'

    # Título
    ultima_col = len(empresas) + 2  # Mes + empresas + Total
    ws.cell(row=1, column=1, value='Histórico de consultas por mes').font = Font(
        bold=True, size=14, color=COLOR_PRIMARIO
    )
    ws.cell(
        row=2, column=1,
        value=f"Generado el {timezone.localtime():%d/%m/%Y %H:%M}",
    ).font = Font(italic=True, size=9, color='808080')

    # Encabezado
    fila_enc = 4
    ws.cell(row=fila_enc, column=1, value='Mes')
    for indice, empresa in enumerate(empresas, start=2):
        ws.cell(row=fila_enc, column=indice, value=empresa)
    ws.cell(row=fila_enc, column=ultima_col, value='Total del mes')
    _estilar_encabezado(ws, fila_enc, ultima_col)

    # Filas de datos
    fila = fila_enc + 1
    for mes in meses:
        ws.cell(row=fila, column=1, value=_nombre_mes(mes)).border = _BORDE
        total_mes = 0
        for indice, empresa in enumerate(empresas, start=2):
            cantidad = conteos.get((mes, empresa), 0)
            total_mes += cantidad
            celda = ws.cell(row=fila, column=indice, value=cantidad)
            celda.alignment = Alignment(horizontal='center')
            celda.border = _BORDE
        celda_total = ws.cell(row=fila, column=ultima_col, value=total_mes)
        celda_total.font = Font(bold=True)
        celda_total.alignment = Alignment(horizontal='center')
        celda_total.fill = PatternFill('solid', fgColor=COLOR_TOTALES)
        celda_total.border = _BORDE
        fila += 1

    # Fila TOTAL histórico
    fila_total = fila
    ws.cell(row=fila_total, column=1, value='TOTAL HISTÓRICO').font = Font(bold=True)
    ws.cell(row=fila_total, column=1).fill = PatternFill('solid', fgColor=COLOR_TOTALES)
    ws.cell(row=fila_total, column=1).border = _BORDE
    for indice, empresa in enumerate(empresas, start=2):
        total_empresa = sum(conteos.get((mes, empresa), 0) for mes in meses)
        celda = ws.cell(row=fila_total, column=indice, value=total_empresa)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal='center')
        celda.fill = PatternFill('solid', fgColor=COLOR_TOTALES)
        celda.border = _BORDE
    celda = ws.cell(row=fila_total, column=ultima_col, value=sum(conteos.values()))
    celda.font = Font(bold=True)
    celda.alignment = Alignment(horizontal='center')
    celda.fill = PatternFill('solid', fgColor=COLOR_TOTALES)
    celda.border = _BORDE

    # Fila PROMEDIO MENSUAL
    fila_prom = fila_total + 1
    num_meses = len(meses) or 1
    ws.cell(row=fila_prom, column=1, value='PROMEDIO MENSUAL').font = Font(bold=True)
    ws.cell(row=fila_prom, column=1).fill = PatternFill('solid', fgColor=COLOR_PROMEDIO)
    ws.cell(row=fila_prom, column=1).border = _BORDE
    for indice, empresa in enumerate(empresas, start=2):
        total_empresa = sum(conteos.get((mes, empresa), 0) for mes in meses)
        celda = ws.cell(row=fila_prom, column=indice, value=round(total_empresa / num_meses, 1))
        celda.font = Font(bold=True)
        celda.number_format = '0.0'
        celda.alignment = Alignment(horizontal='center')
        celda.fill = PatternFill('solid', fgColor=COLOR_PROMEDIO)
        celda.border = _BORDE
    celda = ws.cell(row=fila_prom, column=ultima_col, value=round(sum(conteos.values()) / num_meses, 1))
    celda.font = Font(bold=True)
    celda.number_format = '0.0'
    celda.alignment = Alignment(horizontal='center')
    celda.fill = PatternFill('solid', fgColor=COLOR_PROMEDIO)
    celda.border = _BORDE

    # Nota al pie
    ws.cell(
        row=fila_prom + 2, column=1,
        value=(f"El promedio se calcula sobre {num_meses} mes(es) del periodo "
               f"(incluye los meses sin consultas)."),
    ).font = Font(italic=True, size=9, color='808080')

    _ajustar_anchos(ws, [22] + [20] * len(empresas) + [15])
    ws.freeze_panes = ws.cell(row=fila_enc + 1, column=2)
    return ws


def _hoja_por_empresa(wb, meses, empresas, conteos):
    """Hoja 2: una fila por empresa con total, actividad y promedio."""
    ws = wb.create_sheet('Resumen por empresa')

    encabezados = [
        'Empresa', 'Total consultas', 'Promedio mensual',
        'Mes con más consultas', 'Máximo en un mes', 'Meses con actividad',
    ]
    ws.cell(row=1, column=1, value='Resumen por empresa').font = Font(
        bold=True, size=14, color=COLOR_PRIMARIO
    )

    fila_enc = 3
    for indice, titulo in enumerate(encabezados, start=1):
        ws.cell(row=fila_enc, column=indice, value=titulo)
    _estilar_encabezado(ws, fila_enc, len(encabezados))

    num_meses = len(meses) or 1
    fila = fila_enc + 1
    for empresa in empresas:
        por_mes = {mes: conteos.get((mes, empresa), 0) for mes in meses}
        total = sum(por_mes.values())
        activos = sum(1 for cantidad in por_mes.values() if cantidad > 0)
        if total:
            mes_pico = max(por_mes, key=lambda m: por_mes[m])
            pico = por_mes[mes_pico]
            etiqueta_pico = _nombre_mes(mes_pico)
        else:
            pico, etiqueta_pico = 0, '—'

        valores = [
            empresa, total, round(total / num_meses, 1),
            etiqueta_pico, pico, f"{activos} de {num_meses}",
        ]
        for indice, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=indice, value=valor)
            celda.border = _BORDE
            if indice > 1:
                celda.alignment = Alignment(horizontal='center')
            if indice == 3:
                celda.number_format = '0.0'
        fila += 1

    # Fila total
    total_general = sum(conteos.values())
    ws.cell(row=fila, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=fila, column=2, value=total_general).font = Font(bold=True)
    celda_prom = ws.cell(row=fila, column=3, value=round(total_general / num_meses, 1))
    celda_prom.font = Font(bold=True)
    celda_prom.number_format = '0.0'
    for col in range(1, len(encabezados) + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = PatternFill('solid', fgColor=COLOR_TOTALES)
        celda.border = _BORDE
        if col > 1:
            celda.alignment = Alignment(horizontal='center')

    _ajustar_anchos(ws, [32, 16, 18, 22, 18, 20])
    ws.freeze_panes = ws.cell(row=fila_enc + 1, column=1)
    return ws


def construir_libro_historico():
    """
    Construye el Workbook con el histórico completo.
    Retorna (workbook, nombre_archivo).
    """
    meses, empresas, conteos, _ = _recolectar_datos()

    wb = Workbook()

    if not meses:
        ws = wb.active
        ws.title = 'Consultas por mes'
        ws.cell(row=1, column=1, value='Histórico de consultas por mes').font = Font(
            bold=True, size=14, color=COLOR_PRIMARIO
        )
        ws.cell(row=3, column=1, value='Todavía no hay consultas registradas en la plataforma.')
        _ajustar_anchos(ws, [60])
    else:
        _hoja_por_mes(wb, meses, empresas, conteos)
        _hoja_por_empresa(wb, meses, empresas, conteos)

    nombre = f"historico_consultas_{timezone.localtime():%Y%m%d}.xlsx"
    return wb, nombre
